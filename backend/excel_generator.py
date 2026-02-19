from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(resultados, filepath, nombre_inmobiliaria):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_inmobiliaria[:31]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    link_font = Font(name="Calibri", color="0563C1", underline="single", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["#", "Región", "Precio", "Descripción", "Habitaciones", "Baños", "Ascensor", "Link"]
    col_widths = [5, 15, 18, 55, 15, 12, 12, 40]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    for idx, prop in enumerate(resultados, 1):
        row = idx + 1
        ws.row_dimensions[row].height = 40

        ws.cell(row=row, column=1, value=idx).alignment = cell_alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=prop.get("region", "")).alignment = cell_alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=prop.get("precio", "")).alignment = cell_alignment
        ws.cell(row=row, column=3).border = thin_border

        desc_cell = ws.cell(row=row, column=4, value=prop.get("descripcion", ""))
        desc_cell.alignment = cell_alignment
        desc_cell.border = thin_border

        ws.cell(row=row, column=5, value=prop.get("habitaciones", "")).alignment = cell_alignment
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=prop.get("banos", "")).alignment = cell_alignment
        ws.cell(row=row, column=6).border = thin_border

        ws.cell(row=row, column=7, value=prop.get("ascensor", "NO")).alignment = cell_alignment
        ws.cell(row=row, column=7).border = thin_border

        link_url = prop.get("link", "")
        link_cell = ws.cell(row=row, column=8)
        if link_url:
            link_cell.value = "Ver inmueble"
            link_cell.hyperlink = link_url
            link_cell.font = link_font
        else:
            link_cell.value = ""
        link_cell.alignment = cell_alignment
        link_cell.border = thin_border

        alt_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid") if idx % 2 == 0 else None
        if alt_fill:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = alt_fill

    wb.save(filepath)
    print(f"Excel generado: {filepath} ({len(resultados)} propiedades)")
